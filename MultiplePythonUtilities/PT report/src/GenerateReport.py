
import pandas as pd
from openpyxl import load_workbook

class reportGeneration:
    def __init__(self,baselineRowfile,targetRowfile,finalReportfile,baselineReportsheet,targetReportSheet,finalReportSheet,):
        """
        Initializes the reportGeneration class with the provided file paths and sheet names.

        Parameters:
        - baselineRowfile (str): File path of the baseline row data.
        - targetRowfile (str): File path of the target row data.
        - finalReportfile (str): File path of the final report.
        - baselineReportsheet (str): Name of the sheet in the final report for baseline data.
        - targetReportSheet (str): Name of the sheet in the final report for target data.
        - finalReportSheet (str): Name of the sheet in the final report for comparison data.
        """
        self.baselineRowfile = baselineRowfile
        self.targetRowfile = targetRowfile
        self.finalReportfile = finalReportfile
        self.baselineReportsheet = baselineReportsheet
        self.targetReportSheet = targetReportSheet
        self.finalReportSheet = finalReportSheet

    def createdataframe(self,baselineRowfile ):
        """
        Reads the baseline row data from a CSV file and performs data transformations.

        Parameters:
        - baselineRowfile (str): File path of the baseline row data.

        Returns:
        - SelectedData (DataFrame): Transformed dataframe containing selected columns and calculated metrics.
        """
        try:
            source = pd.read_csv(baselineRowfile)
            SelectedData = source[['Label','Min','Average','90% Line','Max','# Samples','failure']]
            # Create 'Pass' column
            SelectedData['Pass'] = SelectedData['# Samples'] - SelectedData['failure']
            # Create 'Fail' column
            SelectedData['Fail'] = SelectedData['failure']
            # Create 'Fail%' column
            SelectedData['Fail%'] = ((SelectedData['failure'] / SelectedData['# Samples']) * 100).round(2)
            SelectedData=SelectedData.drop(['# Samples','failure'],axis=1)
            SelectedData = SelectedData[[   'Label', 'Min', 'Average', '90% Line', 'Max', 'Pass', 'Fail', 'Fail%']] 
            # Filter rows that start with 'TC' in 'Label' column
            SelectedData = SelectedData[SelectedData['Label'].str.startswith('TC')]
            print(SelectedData['Pass'].sum())
            print(SelectedData['Fail'].sum())
            return SelectedData
        except Exception as e:
            print(e)
            return None

    def updateCell(self,finalReportfile,finalReportSheet,cellname,value):
        """
        Updates a cell in the final report with the provided value.

        Parameters:
        - finalReportfile (str): File path of the final report.
        - finalReportSheet (str): Name of the sheet in the final report.
        - cellname (str): Name of the cell to be updated.
        - value (str): Value to be written in the cell.
        """
        try:
            # Load existing workbook
            self.book = load_workbook(finalReportfile, data_only=True)
            self.book[finalReportSheet][cellname] = value
            self.book.save(finalReportfile)      
            print("Data written to Excel file")
        except Exception as e:
            print(e)

    def updateReportSheet(self,dataFrame,finalReportfile,baselineReportsheet,startrow1,startcol1):   
        """
        Updates a sheet in the final report with the provided dataframe.

        Parameters:
        - dataFrame (DataFrame): Dataframe to be written to the sheet.
        - finalReportfile (str): File path of the final report.
        - baselineReportsheet (str): Name of the sheet in the final report.
        - startrow1 (int): Starting row index to write the dataframe.
        - startcol1 (int): Starting column index to write the dataframe.
        """
        try:
            # Load existing workbook
            self.book = load_workbook(finalReportfile, data_only=True)
            # Create writer with append mode and loaded workbook
            with pd.ExcelWriter(finalReportfile, engine='openpyxl', mode='a',if_sheet_exists='overlay') as writer:
                # Write data to excel sheet   
                dataFrame.to_excel(writer, sheet_name=baselineReportsheet,startrow=startrow1,startcol=startcol1, header=False, index=False)  # Or append to existing sheet

            print("Data written to Excel file")
        except Exception as e:
            print(e)

    def combineDataframes(self,dataFrame1,dataFrame2,listofcolumnstobeDroped):
        """
        Combines two dataframes and performs calculations to create a new dataframe.

        Parameters:
        - dataFrame1 (DataFrame): First dataframe to be combined.
        - dataFrame2 (DataFrame): Second dataframe to be combined.
        - listofcolumnstobeDroped (list): List of columns to be dropped from the combined dataframe.

        Returns:
        - combinedDataFrame (DataFrame): Combined dataframe with calculated metrics.
        """
        try:
            dataFrame1= dataFrame1.reset_index(drop=True)
            dataFrame2= dataFrame2.reset_index(drop=True)
            combinedDataFrame = pd.concat([dataFrame1,dataFrame2 ], axis=1)            
            combinedDataFrame['AVG'] =combinedDataFrame[['target_Average','base_Average']].mean(axis=1)
            combinedDataFrame['90%'] =combinedDataFrame[['target_90% Line','base_90% Line']].mean(axis=1)           
            combinedDataFrame = combinedDataFrame.drop(listofcolumnstobeDroped,axis=1)
            combinedDataFrame.to_csv('dataframe.txt', sep='\t', index=False)
            print(combinedDataFrame)
            return combinedDataFrame

        except Exception as e:
            print(e)
            return None

# Create object of class reportGeneration
baselineRowfile = 'PT report/inputfile/InputData1.csv'
targetRowfile = 'PT report/inputfile/InputData2.csv'
finalReportfile = 'PT report/Outputfile/Report.xlsx'
baselineReportsheet = 'Baseline'
targetReportSheet = 'Target'
finalReportSheet = 'Comparison'
objectReportGeneration = reportGeneration(baselineRowfile,targetRowfile,finalReportfile,baselineReportsheet,targetReportSheet,finalReportSheet)
objectReportGeneration.updateReportSheet(objectReportGeneration.createdataframe(baselineRowfile),finalReportfile,baselineReportsheet,22,0)
objectReportGeneration.updateReportSheet(objectReportGeneration.createdataframe(targetRowfile),finalReportfile,targetReportSheet,22,0)
updatedbaselineDf = objectReportGeneration.createdataframe(baselineRowfile)

updatedbaselineDf.columns = ['base_' + col for col in updatedbaselineDf.columns]
updatedtargetDf = objectReportGeneration.createdataframe(targetRowfile)
updatedtargetDf.columns = ['target_' +col for col in updatedtargetDf.columns]
objectReportGeneration.updateReportSheet(objectReportGeneration.combineDataframes(updatedbaselineDf,updatedtargetDf,['target_Fail%','base_Fail%','target_Label']),finalReportfile,finalReportSheet,5,0)
objectReportGeneration.updateCell(finalReportfile,baselineReportsheet,'H13',objectReportGeneration.createdataframe(baselineRowfile)['Pass'].sum())
objectReportGeneration.updateCell(finalReportfile,baselineReportsheet,'H14',objectReportGeneration.createdataframe(baselineRowfile)['Fail'].sum())